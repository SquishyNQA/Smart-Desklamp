import cv2
from simple_facerec import SimpleFacerec
import time
import threading
import keyboard
import serial.tools.list_ports
import sys
import openpyxl
import queue
from datetime import datetime


sfr = SimpleFacerec()
sfr.load_encoding_images("images/")

cap = cv2.VideoCapture(0)

start_time = None
end_time = None
time_passed = None
check = True
on_check = False
turn_off_no = 0
stop_flag = False

def return_time(start_time):
    current_datetime = datetime.fromtimestamp(start_time)
    formatted_time = current_datetime.strftime("%H:%M:%S")
    date = current_datetime.strftime("%A")
    return formatted_time, date

def camera():
    while True:
        if stop_flag:
            print("CAMERA STOPPED")
            return
        else:
            ret, frame = cap.read()
            count = 0
            if ret:
                if count % 5 == 0:
                    #print(frame)
                    face_locations, face_names = sfr.detect_known_faces(frame)
                    for face_loc, name in zip(face_locations, face_names):
                    # y1, x2, y2, x1 = face_loc[0],face_loc[1],face_loc[2],face_loc[3]
                        if name == "Calvin":
                            yield name
                                # Start the stopwatch when "Calvin" is enteredưq
                            # Small delay to avoid immediate stop
                            #cv2.putText(frame, name, (x1, y1-10), cv2.FONT_HERSHEY_COMPLEX, 1, (0,0,200), 2)
                            #cv2.rectangle(frame, (x1, y1), (x2, y2), (0,0,200), 4)
                                #if 'Calvin' in name:
                                    #start_event.set()
                        else:
                            yield None       
                            #print(face_loc)
                    #cv2.imshow("Frame", frame)
            else:
                print("ERRORRRR")
                cap.release()
                cv2.destroyAllWindows()
            count += 1


def arduino_input(input):
    if input == "OFF":
        #time.sleep(0.05)
        serialInst.write(input.encode('utf-8'))
        print("STOP")
        time.sleep(1)
    elif input == "ON":
        serialInst.write(input.encode('utf-8'))
        print("ON")
        time.sleep(1)
    elif input == "FLICKER":
        serialInst.write(input.encode('utf-8'))
        print("FLICKER")
        time.sleep(1)


def stopwatch(start_event, stop_event):
    global check
    global on_check
    global turn_off_no 
    global stop_flag
    while True:
        if stop_flag:
            print("STOPWATCH STOP")
            return
        else:
            start_event.wait()
            #start_event.clear()
            #print("BRUHH")
            time.sleep(0.1)

            if start_time != None:
                time_has_passed_since_Calvin = time.time() - start_time
                if time_has_passed_since_Calvin > 90:
                    time.sleep(0.05)
                    if check:
                        arduino_input("OFF")
                        time_off,donotneed = return_time(time.time())
                        sheet.cell(row = max_row + turn_off_no + 1, column = 3, value = time_off)
                        #total_time = sheet.cell(row = max_row + turn_off_no + 1, column = 2) - time_off
                        #sheet.cell(row = max_row + turn_off_no + 1, column = 4, value = total_time)
                        workbook.save('light data.xlsx')
                        turn_off_no +=1
                        check = False
                        on_check = False
                elif time_has_passed_since_Calvin <= 90 and time_has_passed_since_Calvin > 75:
                    time.sleep(0.05)
                    if check:
                        arduino_input("FLICKER")
                        #time.sleep(0.05)
                    #print("HELLO????")
            
            time.sleep(0.1)

def writing_data(start):
    
    final_time, weekday = return_time(start)
    sheet.cell(row = max_row + turn_off_no + 1, column = 1, value = max_row + turn_off_no -1)
    sheet.cell(row = max_row + turn_off_no + 1, column = 2, value = final_time)
    sheet.cell(row = max_row + turn_off_no + 1, column = 4, value = weekday)
    workbook.save('light data.xlsx')

def task(start_event):
    calvin_count = 1
    global check
    global stop_flag
    global on_check
    global start_time
    global end_time
    global time_passed
    global turn_off_no 
    decision_generator = camera()
    print("Camera has started")
    while True:

        if stop_flag:
            print("CAMERA STOPPED")
            return
        else:
            decision = next(decision_generator)
            #q
            #print(decision)
            if "Calvin" in str(decision):
                if calvin_count == 1:
                    start_time = time.time()
                    print("Stopwatch has started")
                    if on_check == False:
                        arduino_input("ON")
                        writing_data(start_time)
                        on_check = True
                    #print(f"{on_check} calvin1")
                    check =  True
                    start_event.set()
                else:
                    #print("IT IS STUPID")
                    #reset_event.set()
                    end_time = time.time()
                    time_passed = round(end_time - start_time,1)
                    start_time = end_time
                    print(f"It has been {time_passed}s since the last Calvin")
                    check =  True
                    if on_check == False:
                        arduino_input("ON")
                        writing_data(start_time)
                        on_check = True
                    #print(on_check)
                    start_event.set()
                #print("Yes")
                calvin_count += 1 
            else:

                print("No")
                #arduino_input("Strange Person")  # Adjust sleep time as needed
            time.sleep(0.1)
        
def close_threads():
    global stop_flag
    stop_flag = True

if __name__ == "__main__":
    workbook = openpyxl.load_workbook('light data.xlsx')
    sheet = workbook['Sheet1']  
    max_row = sheet.max_row
    print(max_row)
    #sheet.cell(row = max_row + 1, column = 1, value = max_row + 1)
    
    ports = serial.tools.list_ports.comports()
    serialInst = serial.Serial()

    portsList = []

    for port in ports:
        portsList.append(str(port))
        print(str(port))

    #val = input("Select port: COM")
    for x in range(0,len(portsList)):
        if portsList[x].startswith("COM3"):
            portvar = "COM3"
            print(f"Arduino was detected. Automatically select {portvar}")

    serialInst.baudrate = 9600
    serialInst.port = portvar
    serialInst.open()
      
    print("It is getting ready")
    time.sleep(5)
    arduino_input("OFF")

    start_event = threading.Event()
    stop_event = threading.Event()
    stop_flag = False
    
    stopwatch_thread = threading.Thread(target=stopwatch, args=(start_event, stop_event))
    camera_thread = threading.Thread(target = task, args=(start_event))
    stopwatch_thread.start()
    camera_thread.start()

    while True:
        if keyboard.is_pressed('F12'):
            arduino_input("OFF")
            close_threads()
            print("Exiting loop...")
            sys.exit(0)

            

